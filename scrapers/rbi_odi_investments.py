"""
RBI Overseas Direct Investment (ODI) scraper.

Source: https://rbi.org.in/Scripts/Data_Overseas_Investment.aspx

Each month RBI publishes a press release with an Excel file listing every
Indian party that made an overseas investment (Equity / Loan / Guarantee in
USD million), the JV/WOS name, country, and major activity.

The listing page uses an ASP.NET postback for year filters - we replay it
with the captured __VIEWSTATE / __EVENTVALIDATION fields. Each year returns
~12 press releases (prid). Each press release page links to a single .xlsx
(2018+) or .xls (2017) file on rbidocs.rbi.org.in.

Excel column layout (same across all years):
    A=Sr  B=IndianParty  C=JV/WOS_name  D=JV_or_WOS  E=Country  F=(empty)
    G=Activity  H=Equity  I=Loan  J=Guarantee  K=Total

This is investment data, NOT a sanctions/watchlist. It goes into its own
table: rbi_odi_investments (not watchlist_records).

Outputs:
    data/rbi_odi/                          - cached Excel files
    data/rbi_odi_investments_master.csv    - combined dataset
"""

import csv
import json
import os
import re
import time
from datetime import datetime

import openpyxl
import requests
import xlrd
from bs4 import BeautifulSoup

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CACHE_DIR = os.path.join(PROJECT_ROOT, "data", "rbi_odi")
MASTER_CSV = os.path.join(PROJECT_ROOT, "data", "rbi_odi_investments_master.csv")
PRID_MANIFEST = os.path.join(CACHE_DIR, "_prid_manifest.json")

LISTING_URL = "https://rbi.org.in/Scripts/Data_Overseas_Investment.aspx"
PRESS_URL = "https://rbi.org.in/Scripts/BS_PressReleaseDisplay.aspx?prid={prid}"

YEARS = [str(y) for y in range(2026, 2010, -1)]  # 2017..2026

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

MASTER_FIELDS = [
    "prid", "press_release_url", "excel_url", "excel_filename",
    "period", "period_from", "period_to",
    "sr_no", "indian_party", "jv_wos_name", "jv_or_wos",
    "country", "activity",
    "equity_usd_mn", "loan_usd_mn", "guarantee_usd_mn", "total_usd_mn",
    "scraped_at",
]


def make_session():
    s = requests.Session()
    s.headers.update(HEADERS)
    # Warm up - cookies from rbi.org.in cover rbidocs.rbi.org.in downloads
    s.get("https://rbi.org.in/", timeout=30)
    return s


def get_listing_prids(session, year):
    """Return list of (prid, link_text) for the given year (POST the ASP.NET form)."""
    r = session.get(LISTING_URL, timeout=30)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")

    form_data = {}
    for inp in soup.find_all("input", type="hidden"):
        name = inp.get("name")
        if name:
            form_data[name] = inp.get("value", "")
    form_data["hdnYear"] = year
    form_data["UsrFontCntr$btn"] = ""

    r2 = session.post(LISTING_URL, data=form_data, timeout=30)
    r2.raise_for_status()
    soup2 = BeautifulSoup(r2.text, "html.parser")

    out = []
    seen = set()
    for a in soup2.find_all("a", href=True):
        m = re.search(r"prid=(\d+)", a["href"])
        if not m:
            continue
        text = a.get_text(strip=True)
        text_l = text.lower()
        # Accept both "Overseas Direct Investment" and "Outward Foreign Direct
        # Investment" - RBI used the latter for a handful of releases (e.g.
        # Dec 2016, Dec 2017, Feb 2018).
        if "overseas direct investment" not in text_l \
                and "outward foreign direct investment" not in text_l:
            continue
        prid = m.group(1)
        if prid in seen:
            continue
        seen.add(prid)
        out.append((prid, text))
    return out


def get_excel_url(session, prid):
    """Fetch press release page; return Excel URL (or None)."""
    url = PRESS_URL.format(prid=prid)
    r = session.get(url, timeout=30)
    if r.status_code != 200:
        return None
    soup = BeautifulSoup(r.text, "html.parser")

    # Look for any link whose href ends with .xlsx or .xls
    candidates = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        text = a.get_text(strip=True).lower()
        href_lower = href.lower()
        if href_lower.endswith(".xlsx") or href_lower.endswith(".xls"):
            candidates.append((text, href))

    # Prefer link whose text mentions overseas / OFDI / outward
    for text, href in candidates:
        if any(k in text for k in ("overseas", "ofdi", "outward")):
            return _abs(href)
    if candidates:
        return _abs(candidates[0][1])
    return None


def _abs(href):
    if href.startswith("http://"):
        return "https://" + href[len("http://"):]
    if href.startswith("https://"):
        return href
    return "https://rbidocs.rbi.org.in" + (href if href.startswith("/") else "/" + href)


def download_excel(session, url, prid):
    """Download Excel; return local path or None."""
    fname = url.rsplit("/", 1)[-1].split("?")[0]
    if not fname.lower().endswith((".xls", ".xlsx")):
        fname = f"prid{prid}.xlsx"
    path = os.path.join(CACHE_DIR, fname)
    if os.path.exists(path) and os.path.getsize(path) > 2000:
        return path

    try:
        r = session.get(
            url, timeout=120,
            headers={**HEADERS, "Referer": PRESS_URL.format(prid=prid)},
            allow_redirects=True,
        )
    except requests.RequestException as e:
        print(f"  ! download error for prid={prid}: {e}")
        return None

    if r.status_code != 200 or len(r.content) < 2000:
        print(f"  ! bad download prid={prid} status={r.status_code} bytes={len(r.content)}")
        return None
    # Reject HTML (bot wall)
    head = r.content[:80].lower()
    if b"<html" in head or b"<!doctype" in head:
        print(f"  ! got HTML instead of Excel for prid={prid} ({len(r.content)} bytes)")
        return None
    with open(path, "wb") as f:
        f.write(r.content)
    return path


# --- Parsing -----------------------------------------------------------------

def _safe_float(v):
    import math as _m
    if v is None or v == "":
        return 0.0
    try:
        f = float(v)
    except (TypeError, ValueError):
        s = str(v).replace(",", "").strip()
        try:
            f = float(s)
        except ValueError:
            return 0.0
    if _m.isnan(f) or _m.isinf(f):
        return 0.0
    return f


def _safe_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _clean(s):
    if s is None:
        return ""
    return str(s).replace("\xa0", " ").strip()


def _parse_period(text):
    """Extract (period_from, period_to). Handles both 'From : DD/MM/YYYY ... To : DD/MM/YYYY'
    (modern files) and 'FROM DD/MM/YYYY TO DD/MM/YYYY' (older cumulative files, no colons)."""
    m = re.search(
        r"From\s*:?\s*(\d{2}/\d{2}/\d{4}).*?To\s*:?\s*(\d{2}/\d{2}/\d{4})",
        text, re.IGNORECASE | re.DOTALL,
    )
    if m:
        return m.group(1), m.group(2)
    return None, None


def _detect_columns(header_row, sub_header_row=None):
    """Map logical field names to column indices using header (and optional
    sub-header for the Equity/Loan/Guarantee/Total row).

    Returns a dict like {'sr': 1, 'indian_party': 2, ...}. Some old files
    have a leading empty column (Sr at col 1 instead of col 0), or no
    'merged' empty column between Country and Activity. Detecting from
    the actual header text handles both."""
    col_map = {}

    def _set(rows, key, predicate):
        for row in rows:
            for i, cell in enumerate(row):
                if cell is None:
                    continue
                t = str(cell).strip().lower()
                if predicate(t):
                    col_map.setdefault(key, i)
                    return

    rows = [header_row]
    if sub_header_row is not None:
        rows.append(sub_header_row)

    _set(rows, "sr",
         lambda t: t in ("sr.", "sr", "s.no.", "s no", "s. no", "s.no")
                    or (t.startswith("sr") and ("no" in t or t == "sr.")))
    _set(rows, "indian_party",
         lambda t: "indian party" in t or "applicant" in t)
    _set(rows, "jv_wos_name",
         lambda t: ("jv/wos" in t or "name of the jv" in t or "name of jv" in t)
                    and "whether" not in t)
    _set(rows, "jv_or_wos",
         lambda t: "whether" in t and ("jv" in t or "wos" in t))
    _set(rows, "country",
         lambda t: "overseas country" in t or t == "country" or t.endswith("country"))
    _set(rows, "activity",
         lambda t: "major activity" in t or t == "activity")
    _set(rows, "equity",
         lambda t: t.startswith("equity"))
    _set(rows, "loan",
         lambda t: t.startswith("loan"))
    _set(rows, "guarantee",
         lambda t: t.startswith("guarantee"))
    _set(rows, "total",
         lambda t: t == "total" or t == "total*" or t.startswith("total "))

    return col_map


def _rows_from_sheet(rows_iter):
    """
    Find the period and header row, then yield parsed data dicts.

    Columns are detected dynamically from the header text. Common layouts:
      Modern (2017+): Sr=0, IndianParty=1, JV/WOS=2, Whether=3, Country=4,
                      (empty=5), Activity=6, Equity=7, Loan=8, Guarantee=9, Total=10
      Older (2011-2015): same fields shifted by one (Sr=1, IndianParty=2, ...)
                         because col 0 is blank
    """
    all_rows = list(rows_iter)
    period_from = period_to = None
    for row in all_rows[:20]:
        for cell in row:
            if cell is None:
                continue
            s = str(cell)
            if "from" in s.lower() and "to" in s.lower():
                pf, pt = _parse_period(s)
                if pf:
                    period_from, period_to = pf, pt
                    break
        if period_from:
            break

    # Find header row containing "Indian Party"
    header_idx = None
    for i, row in enumerate(all_rows):
        for cell in row:
            if cell and "indian party" in str(cell).lower():
                header_idx = i
                break
        if header_idx is not None:
            break
    if header_idx is None:
        return

    header_row = all_rows[header_idx]
    sub_header_row = all_rows[header_idx + 1] if header_idx + 1 < len(all_rows) else None
    col_map = _detect_columns(header_row, sub_header_row)

    if "indian_party" not in col_map:
        return

    # Data starts after the header. If the next row has "Equity" anywhere,
    # it's the sub-header row - skip it. Then skip any blank rows.
    data_start = header_idx + 1
    if data_start < len(all_rows):
        nxt = all_rows[data_start]
        if any(c and "equity" in str(c).lower() for c in nxt if c is not None):
            data_start += 1
    sr_col = col_map.get("sr", 0)
    while data_start < len(all_rows):
        r = all_rows[data_start]
        if r and sr_col < len(r) and r[sr_col] is not None and str(r[sr_col]).strip():
            break
        data_start += 1

    ip_col = col_map["indian_party"]
    jv_col = col_map.get("jv_wos_name", ip_col + 1)
    wh_col = col_map.get("jv_or_wos", jv_col + 1)
    cn_col = col_map.get("country", wh_col + 1)
    act_col = col_map.get("activity", cn_col + 2)
    eq_col = col_map.get("equity", act_col + 1)
    ln_col = col_map.get("loan", eq_col + 1)
    gu_col = col_map.get("guarantee", ln_col + 1)
    tot_col = col_map.get("total", gu_col + 1)

    def _cell(row, idx):
        return row[idx] if 0 <= idx < len(row) else None

    for row in all_rows[data_start:]:
        if not row:
            continue
        sr = _cell(row, sr_col)
        if sr is None or str(sr).strip() == "":
            continue
        sr_int = _safe_int(sr)
        if sr_int is None:
            label = _clean(sr).lower()
            if "total" in label or "note" in label or "*" in label:
                break
            continue

        indian_party = _clean(_cell(row, ip_col))
        if not indian_party or len(indian_party) < 2:
            continue

        yield {
            "period_from": period_from,
            "period_to": period_to,
            "sr_no": sr_int,
            "indian_party": indian_party,
            "jv_wos_name": _clean(_cell(row, jv_col)),
            "jv_or_wos": _clean(_cell(row, wh_col)),
            "country": _clean(_cell(row, cn_col)),
            "activity": _clean(_cell(row, act_col)),
            "equity_usd_mn": _safe_float(_cell(row, eq_col)),
            "loan_usd_mn": _safe_float(_cell(row, ln_col)),
            "guarantee_usd_mn": _safe_float(_cell(row, gu_col)),
            "total_usd_mn": _safe_float(_cell(row, tot_col)),
        }


def parse_excel(path):
    """Parse .xlsx (openpyxl) or .xls (xlrd). Return (first_period_label, rows).
    Each row carries its OWN sheet name in the 'period' key, so cumulative
    historical files (multi-sheet) are handled correctly."""
    out = []
    first_label = None
    is_xls = path.lower().endswith(".xls") and not path.lower().endswith(".xlsx")
    try:
        if is_xls:
            wb = xlrd.open_workbook(path)
            for sname in wb.sheet_names():
                if "summary" in sname.lower():
                    continue
                ws = wb.sheet_by_name(sname)
                rows = (
                    tuple(ws.cell_value(i, j) for j in range(ws.ncols))
                    for i in range(ws.nrows)
                )
                if first_label is None:
                    first_label = sname
                for r in _rows_from_sheet(rows):
                    r["period"] = sname
                    out.append(r)
        else:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                for sname in wb.sheetnames:
                    if "summary" in sname.lower():
                        continue
                    ws = wb[sname]
                    if first_label is None:
                        first_label = sname
                    for r in _rows_from_sheet(ws.iter_rows(values_only=True)):
                        r["period"] = sname
                        out.append(r)
            finally:
                wb.close()
    except Exception as e:
        print(f"  ! parse error {os.path.basename(path)}: {e}")
        return None, []
    return first_label, out


# --- Orchestration -----------------------------------------------------------

def build_manifest(session, force=False):
    """Collect all (prid, year, title) tuples by paging through the year filter."""
    if os.path.exists(PRID_MANIFEST) and not force:
        with open(PRID_MANIFEST) as f:
            return json.load(f)

    manifest = []
    seen = set()
    for year in YEARS:
        try:
            items = get_listing_prids(session, year)
        except Exception as e:
            print(f"  ! year {year} listing failed: {e}")
            continue
        new = 0
        for prid, text in items:
            if prid in seen:
                continue
            seen.add(prid)
            manifest.append({"prid": prid, "year": year, "title": text})
            new += 1
        print(f"  year {year}: {len(items)} links, {new} new (total {len(manifest)})")
        time.sleep(1)

    with open(PRID_MANIFEST, "w") as f:
        json.dump(manifest, f, indent=2)
    return manifest


def main():
    os.makedirs(CACHE_DIR, exist_ok=True)
    session = make_session()

    print("[1/4] Collecting press release manifest across years 2017-2026...")
    manifest = build_manifest(session, force=True)
    print(f"  -> {len(manifest)} ODI press releases\n")

    print("[2/4] Resolving Excel URLs and downloading...")
    excel_records = []  # list of (entry, excel_url, local_path)
    for i, entry in enumerate(manifest, 1):
        prid = entry["prid"]
        excel_url = get_excel_url(session, prid)
        if not excel_url:
            print(f"  [{i}/{len(manifest)}] prid={prid}: no excel link found - skip")
            time.sleep(0.5)
            continue
        path = download_excel(session, excel_url, prid)
        if not path:
            time.sleep(0.5)
            continue
        excel_records.append((entry, excel_url, path))
        if i % 10 == 0 or i == len(manifest):
            print(f"  [{i}/{len(manifest)}] prid={prid} -> {os.path.basename(path)}")
        time.sleep(1)
    print(f"  -> {len(excel_records)} Excel files cached\n")

    print("[3/4] Parsing all Excel files...")
    all_rows = []
    scraped_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    for entry, excel_url, path in excel_records:
        first_label, rows = parse_excel(path)
        for r in rows:
            r["prid"] = entry["prid"]
            r["press_release_url"] = PRESS_URL.format(prid=entry["prid"])
            r["excel_url"] = excel_url
            r["excel_filename"] = os.path.basename(path)
            # parse_excel already set r["period"] to the sheet's own name.
            # Only fall back to the press release title if parsing produced no period.
            if not r.get("period"):
                r["period"] = first_label or entry["title"]
            r["scraped_at"] = scraped_at
            all_rows.append(r)
        if len(rows) == 0:
            print(f"  ! 0 rows from {os.path.basename(path)} (prid={entry['prid']})")
    print(f"  -> {len(all_rows):,} total investment records\n")

    print("[4/4] Writing master CSV...")
    with open(MASTER_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=MASTER_FIELDS, extrasaction="ignore")
        w.writeheader()
        for r in all_rows:
            w.writerow(r)
    print(f"  -> {MASTER_CSV} ({len(all_rows):,} rows)")

    # Summary
    if all_rows:
        parties = set(r["indian_party"] for r in all_rows)
        countries = set(r["country"] for r in all_rows if r["country"])
        total_usd = sum(r["total_usd_mn"] for r in all_rows)
        print()
        print(f"  Unique Indian parties: {len(parties):,}")
        print(f"  Unique countries:      {len(countries)}")
        print(f"  Total USD mn:          {total_usd:,.2f}")


if __name__ == "__main__":
    main()
